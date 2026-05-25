"""FR 6.0 输出导出引擎。

本模块只负责把已经生成/审查过的测试工件转换为标准文件格式，不直接依赖
FastAPI 路由层。这样前端、API 路由和后续批处理脚本都可以复用同一套导出逻辑。
"""

from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timezone
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from app.models.requirement import StructuredRequirement
from app.models.test_case import CoverageItem, TestCase, TestSuite

ExportFormat = Literal["json", "csv", "xlsx"]


class ExportOptions(BaseModel):
    """前端传入的导出配置。

    include_* 字段让演示时可以清楚说明导出范围；默认全部包含，符合 FR 6.0
    对测试用例、测试套件、风险评分和覆盖信息的交付要求。
    """

    include_requirements: bool = True
    include_test_cases: bool = True
    include_coverage: bool = True
    include_summary: bool = True
    file_basename: str = Field(default="intellitest_export", min_length=1)


class ExportArtifact(BaseModel):
    """统一导出数据包。

    成员 A/B/C/D 的输出最终会汇集成这个结构：需求与风险来自 A/B，测试用例与
    覆盖项来自 C/D，成员 E 负责把这些内容标准化输出。
    """

    suite: TestSuite | None = None
    requirements: list[StructuredRequirement] = Field(default_factory=list)
    test_cases: list[TestCase] = Field(default_factory=list)
    coverage_items: list[CoverageItem] = Field(default_factory=list)
    options: ExportOptions = Field(default_factory=ExportOptions)


class ExportResult(BaseModel):
    """导出结果元数据，供 API 路由设置响应头。"""

    content: bytes
    media_type: str
    filename: str


def export_artifact(artifact: ExportArtifact, file_format: ExportFormat) -> ExportResult:
    """根据目标格式导出测试工件。

    Args:
        artifact: 已汇总的测试工件数据。
        file_format: 目标格式，支持 json/csv/xlsx。

    Returns:
        含文件二进制内容、MIME 类型和建议文件名的导出结果。
    """

    basename = _safe_basename(artifact.options.file_basename)
    if file_format == "json":
        return ExportResult(
            content=_build_json(artifact),
            media_type="application/json",
            filename=f"{basename}.json",
        )
    if file_format == "csv":
        return ExportResult(
            content=_build_csv(artifact),
            media_type="text/csv; charset=utf-8",
            filename=f"{basename}.csv",
        )
    if file_format == "xlsx":
        return ExportResult(
            content=_build_xlsx(artifact),
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            filename=f"{basename}.xlsx",
        )
    raise ValueError(f"Unsupported export format: {file_format}")


def _collect_cases(artifact: ExportArtifact) -> list[TestCase]:
    """合并 suite 内用例与独立用例，并按 ID 去重。"""

    cases: dict[str, TestCase] = {}
    if artifact.suite:
        for case in artifact.suite.test_cases:
            cases[case.id] = case
    for case in artifact.test_cases:
        cases[case.id] = case
    return list(cases.values())


def _build_json(artifact: ExportArtifact) -> bytes:
    options = artifact.options
    payload = {
        "schema": "IntelliTest.FR6.Export.v1",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "suite": _suite_to_json(artifact),
        "requirements": [
            item.model_dump(mode="json") for item in artifact.requirements
        ]
        if options.include_requirements
        else [],
        "test_cases": [
            _case_to_json(item, include_coverage=options.include_coverage)
            for item in _collect_cases(artifact)
        ]
        if options.include_test_cases
        else [],
        "coverage_items": [
            item.model_dump(mode="json") for item in artifact.coverage_items
        ]
        if options.include_coverage
        else [],
        "summary": _build_summary(artifact) if options.include_summary else None,
        "options": options.model_dump(mode="json"),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def _suite_to_json(artifact: ExportArtifact) -> dict[str, object] | None:
    if artifact.suite is None:
        return None

    options = artifact.options
    payload = artifact.suite.model_dump(mode="json")
    payload["test_cases"] = [
        _case_to_json(item, include_coverage=options.include_coverage)
        for item in artifact.suite.test_cases
    ] if options.include_test_cases else []
    return payload


def _case_to_json(case: TestCase, include_coverage: bool) -> dict[str, object]:
    payload = case.model_dump(mode="json")
    if not include_coverage:
        payload.pop("coverage_items", None)
    return payload


def _build_csv(artifact: ExportArtifact) -> bytes:
    """生成适合测试管理工具导入的扁平 CSV。

    CSV 以测试用例为主表，每行保留需求、风险、技术和覆盖项追溯信息。
    """

    stream = io.StringIO(newline="")
    requirement_titles = _requirement_title_map(artifact)
    fieldnames = [
        "test_case_id",
        "requirement_id",
        "requirement_title",
        "title",
        "precondition",
        "test_steps",
        "test_data",
        "expected_result",
        "technique",
        "priority",
        "risk_score",
        "modified_by_user",
    ]
    if artifact.options.include_coverage:
        fieldnames.insert(-1, "coverage_items")

    writer = csv.DictWriter(stream, fieldnames=fieldnames)
    writer.writeheader()
    for case in _collect_cases(artifact):
        writer.writerow(
            _case_to_row(
                case,
                include_coverage=artifact.options.include_coverage,
                requirement_titles=requirement_titles,
            )
        )
    return stream.getvalue().encode("utf-8-sig")


def _build_xlsx(artifact: ExportArtifact) -> bytes:
    """生成带样式的 Excel 工作簿，便于演示和人工审查。"""

    wb = Workbook()
    first_sheet = wb.active
    if artifact.options.include_summary:
        first_sheet.title = "Summary"
        _write_summary_sheet(first_sheet, artifact)
    else:
        first_sheet.title = "Export"
        first_sheet.append(["IntelliTest export"])
        _style_sheet(first_sheet)

    if artifact.options.include_test_cases:
        cases_sheet = wb.create_sheet("Test Cases")
        requirement_titles = _requirement_title_map(artifact)
        _write_table(
            cases_sheet,
            [
                _case_to_row(
                    case,
                    include_coverage=artifact.options.include_coverage,
                    requirement_titles=requirement_titles,
                )
                for case in _collect_cases(artifact)
            ],
        )

    if artifact.options.include_requirements:
        req_sheet = wb.create_sheet("Requirements")
        _write_table(req_sheet, [_requirement_to_row(item) for item in artifact.requirements])

    if artifact.options.include_coverage:
        coverage_sheet = wb.create_sheet("Coverage")
        requirement_titles = _requirement_title_map(artifact)
        _write_table(
            coverage_sheet,
            [_coverage_to_row(item, requirement_titles) for item in artifact.coverage_items],
        )

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _write_summary_sheet(sheet, artifact: ExportArtifact) -> None:
    summary = _build_summary(artifact)
    summary_rows = [
        ("Export Schema", "IntelliTest.FR6.Export.v1"),
        ("Generated At", summary["generated_at"]),
        ("Suite ID", summary["suite_id"]),
        ("Suite Name", summary["suite_name"]),
        ("Requirement Count", summary["requirement_count"]),
        ("Test Case Count", summary["test_case_count"]),
        ("Coverage Item Count", summary["coverage_item_count"]),
    ]
    sheet.append(["Item", "Value"])
    for row in summary_rows:
        sheet.append(row)
    _style_sheet(sheet)


def _write_table(sheet, rows: list[dict[str, object]]) -> None:
    if not rows:
        sheet.append(["No data"])
        _style_sheet(sheet)
        return
    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    _style_sheet(sheet)


def _style_sheet(sheet) -> None:
    """统一 Excel 样式，避免导出文件像未经整理的原始数据。"""

    header_fill = PatternFill("solid", fgColor="EAF2FF")
    header_font = Font(bold=True, color="16325C")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_length + 3, 14), 48)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"


def _case_to_row(
    case: TestCase,
    include_coverage: bool = True,
    requirement_titles: dict[str, str] | None = None,
) -> dict[str, object]:
    title_map = requirement_titles or {}
    row: dict[str, object] = {
        "test_case_id": case.id,
        "requirement_id": case.requirement_id,
        "requirement_title": title_map.get(case.requirement_id, ""),
        "title": case.title,
        "precondition": case.precondition or "",
        "test_steps": "\n".join(case.test_steps),
        "test_data": case.test_data or "",
        "expected_result": case.expected_result or "",
        "technique": _plain_value(case.technique),
        "priority": _plain_value(case.priority),
        "risk_score": case.risk_score or "",
        "modified_by_user": "Yes" if case.modified_by_user else "No",
    }
    if include_coverage:
        row["coverage_items"] = ", ".join(case.coverage_items)
    return row


def _requirement_to_row(requirement: StructuredRequirement) -> dict[str, object]:
    return {
        "requirement_id": requirement.id,
        "title": requirement.title or "",
        "raw_text": requirement.raw_text,
        "input_fields": ", ".join(requirement.input_fields),
        "data_ranges": ", ".join(requirement.data_ranges),
        "conditions": ", ".join(requirement.conditions),
        "expected_actions": ", ".join(requirement.expected_actions),
        "risk_score": requirement.risk_score or "",
        "priority": requirement.priority or "",
    }


def _coverage_to_row(
    item: CoverageItem,
    requirement_titles: dict[str, str] | None = None,
) -> dict[str, object]:
    title_map = requirement_titles or {}
    return {
        "coverage_id": item.id,
        "requirement_id": item.requirement_id,
        "requirement_title": title_map.get(item.requirement_id, ""),
        "description": item.description,
        "item_type": item.item_type,
        "selected_techniques": ", ".join(_plain_value(tech) for tech in item.selected_techniques),
        "covered_by_test_cases": ", ".join(item.covered_by_test_cases),
    }


def _safe_basename(name: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in "-_." else "_" for ch in name.strip())
    return cleaned or "intellitest_export"


def _plain_value(value: object) -> str:
    """Return enum values instead of Python enum repr strings for tabular exports."""

    if value is None:
        return ""
    return getattr(value, "value", str(value))


def _requirement_title_map(artifact: ExportArtifact) -> dict[str, str]:
    return {
        requirement.id: requirement.title
        for requirement in artifact.requirements
        if requirement.title
    }


def _build_summary(artifact: ExportArtifact) -> dict[str, object]:
    cases = _collect_cases(artifact)
    options = artifact.options
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "suite_id": artifact.suite.id if artifact.suite else "-",
        "suite_name": artifact.suite.name if artifact.suite else "-",
        "requirement_count": len(artifact.requirements) if options.include_requirements else 0,
        "test_case_count": len(cases) if options.include_test_cases else 0,
        "coverage_item_count": len(artifact.coverage_items) if options.include_coverage else 0,
    }
